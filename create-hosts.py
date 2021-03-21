#!/usr/bin/python3
import os
import re
import openpyxl
wb=openpyxl.load_workbook(filename='esxi-hosts.xlsx')
ws=wb["Hosts"]
iso_folder="/iso/"
http_folder="/var/www/httpboot/"
table=[]

for i in ws.iter_rows(min_row=3):
    str_hostname=i[0].value
    str_nic=i[1].value
    str_ip=i[2].value
    str_netmask=i[3].value
    str_gateway=i[4].value
    str_dns=i[5].value
    str_vlan=i[6].value
    str_idrac_ip=i[7].value
    str_disk_sel=i[8].value
    str_cap_disk=i[9].value
    str_ntp=i[10].value
    str_domain=i[11].value
    str_password=i[12].value
    str_ceip=i[13].value
    str_tps=i[14].value
    str_vsan=i[15].value
    str_esxi=i[16].value
    str_deployment=i[17].value

    values=[str_deployment,str_hostname,str_ip,str_netmask,str_gateway,str_dns,str_vlan,str_idrac_ip,str_domain]
    table+=values
    if not 'deployments' in locals():
        deployments=[str_deployment]
    elif not str_deployment in deployments:
        deployments+=[str_deployment]

    ks='ks={}/{}/ks.cfg'.format(ws['B1'].value,str_hostname)
    ks_folder='{}{}'.format(http_folder,str_hostname)
    #if not os.path.exists(str_hostname):
    #    os.mkdir(str_hostname)
    if not os.path.exists(ks_folder):
        os.mkdir(ks_folder)
    output=open("{}/ks.cfg".format(ks_folder),"w+")

    output.write('vmaccepteula\n')
 
    output.write('rootpw {}\n'.format(str_password))
    output.write('clearpart --alldrives --overwritevmfs\n')

    output.write('install {} --overwritevmfs --novmfsondisk\n'.format(str_disk_sel))
    output.write('keyboard Finnish\n')

    output.write("network --bootproto=static --device={} --ip={} --netmask={} --gateway={} --nameserver={} --hostname={} --vlanid={} --addvmportgroup=1\n".format(str_nic,str_ip,str_netmask,str_gateway,str_dns,str_hostname,int(str_vlan)))

    output.write('reboot --noeject\n')

    output.write('%firstboot --interpreter=busybox\n')
    if str_cap_disk != None:
        capacitydisks=str_cap_disk.split(',')
        for disk in capacitydisks:
            if disk!='':
                output.write('esxcli vsan storage tag add -t capacityFlash -d `esxcli storage core device list|grep -B 1 "Display Name:.*{}"|head -n 1`\n'.format(disk))

    output.write('esxcli network ip dns search add --domain={}\n'.format(str_domain))

    output.write('esxcli network vswitch standard portgroup set -v {} -p "VM Network"\n'.format(int(str_vlan)))

    output.write('vim-cmd hostsvc/enable_ssh\n')

    ntps=str_ntp.split(',')
    for ntp in ntps:
        if ntp!='':
            output.write('echo "server {}" >> /etc/ntp.conf\n'.format(ntp))


    output.write('/sbin/chkconfig ntpd on\n')
    output.write('esxcli system settings advanced set -o /UserVars/HostClientCEIPOptIn -i {}\n'.format(str_ceip))
    if (str_tps == "yes"):
        output.write('esxcli system settings advanced set -o "/Mem/AllocGuestLargePage" --int-value 0\n')
        output.write('esxcli system settings advanced set -o "/Mem/ShareForceSalting" --int-value 0\n')
    if (str_vsan == "yes"):
        output.write('esxcli vsan network ipv4 add -i {}\n'.format(str_nic))
        output.write('esxcli vsan cluster new\n')
        output.write('esxcli vsan policy setdefault -c cluster -p "((\"hostFailuresToTolerate\" i0) (\"forceProvisioning\" i1))"\n')
        output.write('esxcli vsan policy setdefault -c vdisk -p "((\"hostFailuresToTolerate\" i0) (\"forceProvisioning\" i1))")"\n')
        output.write('esxcli vsan policy setdefault -c vmnamespace -p "((\"hostFailuresToTolerate\" i0) (\"forceProvisioning\" i1))"\n')
        output.write('esxcli vsan policy setdefault -c vmswap -p "((\"hostFailuresToTolerate\" i0) (\"forceProvisioning\" i1))"\n')
        output.write('esxcli vsan policy setdefault -c vmem -p "((\"hostFailuresToTolerate\" i0) (\"forceProvisioning\" i1))"\n')
    output.write('reboot\n')
    output.close()
    
    if os.path.isfile("{}{}/boot.cfg".format(iso_folder,str_esxi)):
        bootcfg=open("{}{}/boot.cfg".format(iso_folder,str_esxi),"r").read()
    elif os.path.isfile("{}{}/BOOT.CFG".format(iso_folder,str_esxi)):
        bootcfg=open("{}{}/BOOT.CFG".format(iso_folder,str_esxi),"r").read()
    else:
        os.system('7z x -y -o{}{} {}{}.iso'.format(iso_folder,str_esxi,iso_folder,str_esxi))
        os.system('find {} -type d -exec chmod o+rx {} \;'.format(iso_folder,'{}'))
        os.system("find {}{} -mindepth 3 -exec rename 's/(.*)\/([^\/]*)/$1\/\L$2/' {} \;".format(iso_folder,str_esxi,'{}'))
        os.system("find {}{} -mindepth 2 -exec rename 's/(.*)\/([^\/]*)/$1\/\L$2/' {} \;".format(iso_folder,str_esxi,'{}'))
        os.system("find {}{} -mindepth 1 -exec rename 's/(.*)\/([^\/]*)/$1\/\L$2/' {} \;".format(iso_folder,str_esxi,'{}'))
        bootcfg=open("{}{}/boot.cfg".format(iso_folder,str_esxi),"r").read()

    
    boot=open("{}/boot.cfg".format(ks_folder),"w+")
    newboot=re.sub("/","",bootcfg,flags=re.M)
    newboot=re.sub(r'title=[^\n]*','title=Loading ESXi installer (https://github.com/oizone/esxihttp)',newboot,flags=re.M)
    newboot=re.sub(r'prefix=[^\n]*','prefix={}/{}/'.format(ws['B1'].value,str_esxi),newboot,flags=re.M)
    newboot=re.sub(r'kernelopt=[^\n]*','kernelopt={}'.format(ks),newboot,flags=re.M)
    boot.write(newboot)
    boot.close()


    if os.path.islink('{}/mboot.efi'.format(ks_folder)):
        os.remove('{}/mboot.efi'.format(ks_folder))

    if not os.path.exists('{}{}'.format(http_folder,str_esxi)):
        os.symlink('{}{}'.format(iso_folder,str_esxi),'{}{}'.format(http_folder,str_esxi))

    if os.path.isfile("{}{}/efi/boot/bootx64.efi".format(iso_folder,str_esxi)):
        os.symlink('{}{}/efi/boot/bootx64.efi'.format(iso_folder,str_esxi),'{}/mboot.efi'.format(ks_folder))
    else:
        os.symlink('{}{}/EFI/BOOT/BOOTX64.EFI'.format(iso_folder,str_esxi),'{}/mboot.efi'.format(ks_folder))

    
for i in deployments:
    print(i)
    for o in table:
        print(o)




            
print(table)
print(deployments)
